import re
from dataclasses import dataclass
from typing import List, Optional, Tuple
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


URL = "https://oddscrowd.com/odds-comparison/basketball/leagues/ncaab/bet-types/spread-fullgame"


@dataclass
class TeamRow:
    team: str
    bets: int
    money: int
    spread: str


@dataclass
class OutRow:
    game: str
    spread: str
    side: str
    angle: str
    diff: int


def extract_percents(text: str) -> Optional[Tuple[int, int]]:
    p = re.findall(r"(\d{1,3})\s*%", text)
    if len(p) < 2:
        return None
    return int(p[0]), int(p[1])


def extract_spread(text: str) -> Optional[str]:
    # first +/- number on the line: +1.5 / -10.5 etc.
    m = re.search(r"([+-]\d+(?:\.\d+)?)", text)
    return m.group(1) if m else None


def extract_team_name(text: str) -> Optional[str]:
    """
    Team name is usually before the first %.
    Example: "Arkansas 43% 75% -1.5 -105"
    """
    t = " ".join(text.split())
    idx = t.find("%")
    if idx == -1:
        return None
    left = t[:idx]
    left = re.sub(r"\s+\d{1,3}\s*$", "", left).strip()
    return left if left else None


def parse_team_line(line: str) -> Optional[TeamRow]:
    perc = extract_percents(line)
    spr = extract_spread(line)
    team = extract_team_name(line)
    if not perc or not spr or not team:
        return None
    bets, money = perc
    if bets > 100 or money > 100:
        return None
    return TeamRow(team=team, bets=bets, money=money, spread=spr)


def write_excel(rows: List[OutRow], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Splits"

    headers = ["game", "spread", "side", "angle", "diff"]
    ws.append(headers)

    for c in ws[1]:
        c.font = Font(bold=True)

    for r in rows:
        ws.append([r.game, r.spread, r.side, r.angle, r.diff])

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 70)

    wb.save(path)


def main():
    out_rows: List[OutRow] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1400, "height": 900})

        page.goto(URL, wait_until="domcontentloaded")
        page.wait_for_timeout(5000)

        # Scroll to load all matchups (OddsCrowd lazy loads)
        for _ in range(60):
            page.mouse.wheel(0, 1400)
            page.wait_for_timeout(250)

        previews = page.locator("a:has-text('Game Preview')")
        n = previews.count()
        print("Found previews:", n)

        seen_games = set()

        for i in range(n):
            link = previews.nth(i)

            # Climb ancestors until we find a DIV block that contains BOTH teams.
            # We detect that by requiring >= 4 percent signs (2 teams x 2 percents).
            container = None
            for level in range(1, 12):
                cand = link.locator(f"xpath=ancestor::div[{level}]")
                try:
                    txt = cand.inner_text(timeout=1500)
                except PlaywrightTimeoutError:
                    continue
                if txt.count("%") >= 4:
                    container = cand
                    break

            if container is None:
                continue

            try:
                block_text = container.inner_text(timeout=1500)
            except PlaywrightTimeoutError:
                continue

            # From this block, find two lines that look like:
            # "<Team> <bets%> <money%> <spread> ..."
            lines = [" ".join(ln.split()) for ln in block_text.split("\n") if ln.strip()]
            candidates = []
            for ln in lines:
                if "Game Preview" in ln:
                    continue
                if ln.count("%") >= 2 and re.search(r"[+-]\d+(?:\.\d+)?", ln):
                    candidates.append(ln)

            # Deduplicate while preserving order, keep first 2
            uniq = []
            seen = set()
            for ln in candidates:
                if ln in seen:
                    continue
                seen.add(ln)
                uniq.append(ln)
                if len(uniq) == 2:
                    break

            if len(uniq) < 2:
                continue

            r1 = parse_team_line(uniq[0])
            r2 = parse_team_line(uniq[1])
            if not r1 or not r2:
                continue

            game_key = tuple(sorted([r1.team, r2.team]))
            if game_key in seen_games:
                continue
            seen_games.add(game_key)

            d1 = r1.money - r1.bets
            d2 = r2.money - r2.bets

            if d1 >= d2:
                side = r1.team
                spread = r1.spread
                diff = d1
                angle = f"{r1.money}-{r1.bets}"
            else:
                side = r2.team
                spread = r2.spread
                diff = d2
                angle = f"{r2.money}-{r2.bets}"

            game = f"{r1.team} vs {r2.team}"
            out_rows.append(
                OutRow(game=game, spread=spread, side=side, angle=angle, diff=diff)
            )

        browser.close()

    out_rows.sort(key=lambda r: r.diff, reverse=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = "oddscrowd_ncaab_spread_splits.xlsx"
    write_excel(out_rows, out_file)
    print(f"Saved: {out_file} ({len(out_rows)} games)")


if __name__ == "__main__":
    main()
