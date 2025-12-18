import re
from dataclasses import dataclass
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


URL = "https://oddscrowd.com/odds-comparison/basketball/leagues/ncaab/bet-types/spread-fullgame"


@dataclass
class TeamRow:
    team: str
    bets: int
    money: int
    spread: str


@dataclass
class OutRow:
    game: str
    spread: str
    side: str
    angle: str
    diff: int


def write_excel(rows: List[OutRow], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Splits"

    headers = ["game", "spread", "side", "angle", "diff"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for r in rows:
        ws.append([r.game, r.spread, r.side, r.angle, r.diff])

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 70)

    wb.save(path)


def extract_two_teamrows_from_block(block_text: str) -> Optional[Tuple[TeamRow, TeamRow]]:
    """
    Pull patterns like:
      Arkansas 43% 75% -1.5
      Texas Tech 57% 25% +1.5
    even if the container has extra columns/odds mixed in.
    """
    # Normalize whitespace so regex works even if layout is weird
    t = " ".join(block_text.split())

    # Team name (greedy but bounded by the first percent),
    # then Bets%, Money%, then the opener spread (+/- number)
    pattern = re.compile(
        r"(?P<team>[A-Za-z0-9 .&'’\-\(\)]+?)\s+"
        r"(?P<bets>\d{1,3})%\s+"
        r"(?P<money>\d{1,3})%\s+"
        r"(?P<spread>[+-]\d+(?:\.\d+)?)"
    )

    matches = list(pattern.finditer(t))
    if len(matches) < 2:
        return None

    def to_row(m) -> TeamRow:
        team = m.group("team").strip()
        bets = int(m.group("bets"))
        money = int(m.group("money"))
        spread = m.group("spread")
        return TeamRow(team=team, bets=bets, money=money, spread=spread)

    r1 = to_row(matches[0])
    r2 = to_row(matches[1])
    return r1, r2


def main():
    out_rows: List[OutRow] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1400, "height": 900})

        page.goto(URL, wait_until="domcontentloaded")
        page.wait_for_timeout(5000)

        # scroll to load all games (OddsCrowd lazy loads)
        for _ in range(70):
            page.mouse.wheel(0, 1400)
            page.wait_for_timeout(250)

        previews = page.locator("a:has-text('Game Preview')")
        n = previews.count()
        print("Found previews:", n)

        seen_games = set()

        for i in range(n):
            link = previews.nth(i)

            # climb ancestor divs until we find one that likely holds both teams
            container = None
            container_text = None
            for level in range(1, 14):
                cand = link.locator(f"xpath=ancestor::div[{level}]")
                try:
                    txt = cand.inner_text(timeout=1200)
                except PlaywrightTimeoutError:
                    continue

                # require at least a few % signs; we’ll regex match after
                if txt.count("%") >= 4:
                    container = cand
                    container_text = txt
                    break

            if not container or not container_text:
                continue

            parsed = extract_two_teamrows_from_block(container_text)
            if i == 0:
                print("FIRST BLOCK SAMPLE:", container_text[:500])

            if not parsed:
                continue

            r1, r2 = parsed

            # Deduplicate games
            key = tuple(sorted([r1.team, r2.team]))
            if key in seen_games:
                continue
            seen_games.add(key)

            d1 = r1.money - r1.bets
            d2 = r2.money - r2.bets

            if d1 >= d2:
                side = r1.team
                spread = r1.spread
                diff = d1
                angle = f"{r1.money}-{r1.bets}"
            else:
                side = r2.team
                spread = r2.spread
                diff = d2
                angle = f"{r2.money}-{r2.bets}"

            game = f"{r1.team} vs {r2.team}"
            out_rows.append(OutRow(game=game, spread=spread, side=side, angle=angle, diff=diff))

        browser.close()

    out_rows.sort(key=lambda r: r.diff, reverse=True)

    out_file = "oddscrowd_ncaab_spread_splits.xlsx"
    write_excel(out_rows, out_file)
    print(f"Saved: {out_file} ({len(out_rows)} games)")


if __name__ == "__main__":
    main()
